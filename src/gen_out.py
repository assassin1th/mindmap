import openpyxl
import database
from openpyxl.writer.excel import save_workbook
from openpyxl.utils import get_column_letter

def gen_outputs (db, group_len):
    wb = openpyxl.Workbook ()

    masters = list (map (lambda x: x[0], db.get_column_unique_values ('master_name')))
    
    sheet = wb.create_sheet ('groups')
    gen_groups (masters, group_len, db, sheet)
    gen_groups_by_mindmap_type (db, group_len, wb)
    gen_weight_by_mindmap_type (db, wb)

    save_workbook (wb, '../out/log.xlsx')

def get_key (m, ms):
    for k, b in ms.items ():
        if m == b:
            return k

def gen_weight_by_mindmap_type (db, wb):
    mindmap_types = list (map (lambda x: x[0], db.get_column_unique_values ('mindmap_type')))

    sheet = wb.create_sheet ('weight_by_mindmap_type')
    masters_by_mmt = dict ()
    for m_type in mindmap_types:
        masters_by_mmt[m_type] = set (map (lambda x: x[0], db.get_by_prop ('master_name', 'mindmap_type', m_type)))

    i = 1
    for j in range (len (mindmap_types)):
        res_group = masters_by_mmt[mindmap_types[j]]

        mmt = set ()
        mmt.add(mindmap_types[j])

        print_groups (mmt, res_group, sheet, db, i, len (mindmap_types))
        i += 1

        for k in range (j + 1, len (mindmap_types)):
            res_group = res_group.union (masters_by_mmt[mindmap_types[k]])
            mmt.add (mindmap_types[k])
            print_groups (mmt, res_group, sheet, db, i, len (mindmap_types))
            i += 1
            
def print_groups (mmt, group, sheet, db, row, max_col):
    l_mmt = list (mmt)
    for i in range (len (l_mmt)):
        sheet[get_column_letter (i + 1) + str (row)] = l_mmt[i]

    if (len (group) > 1):
        sheet[get_column_letter (max_col + 1) + str (row)] = group_weight (group, db)

def gen_groups_by_mindmap_type (db, group_len, wb):
    mindmap_types = list (map (lambda x: x[0], db.get_column_unique_values ('mindmap_type')))
    
    for mindmap_type in mindmap_types:
        masters = list (map (lambda x: x[0], db.get_by_prop ('master_name', 'mindmap_type', mindmap_type)))
        
        if len (masters) > 1:
            sheet = wb.create_sheet (mindmap_type)
            gen_groups (masters, group_len, db, sheet)

def group_weight (group, db):
    n = len (group)
    l_group = list (group)
    sum_weight = float (0)
    for first_master in l_group:
        l_group.remove (first_master)

        first_files_set = set (map (lambda x: x[0], db.get_by_prop ('id', 'master_name', first_master)))
        for second_master in l_group:
            second_files_set = set (map (lambda x: x[0], db.get_by_prop ('id', 'master_name', second_master)))
            sum_weight += len (first_files_set.intersection (second_files_set))

    return sum_weight * 2 / (n - 1) / n
    

def gen_groups (masters, group_len, db, sheet):
    for i in range (group_len):
        sheet[get_column_letter (i + 1) + '1'] = 'чел №' + str (i)

    sheet[get_column_letter (group_len + 1) + '1'] = 'общие'
    sheet[get_column_letter (group_len + 2) + '1'] = 'вес группы'

    row = 2
    while len (masters) >= group_len:
        sim_files = set (map (lambda x: x[0], db.get_data_by_prop ('id')))
        group = set ()

        for i in range (group_len):
            max_master = masters[0]
            max_sim_files = set ()

            for master in masters :
                master_files = set (map (lambda x: x[0], db.get_by_prop ('id', 'master_name', master)))
                
                tmp_sim_files = sim_files.intersection (master_files)
                if len (tmp_sim_files) > len (max_sim_files):
                    max_sim_files = tmp_sim_files
                    max_master = master
            if not max_sim_files:
                if i == 0:
                    masters.remove (max_master)
                break
            
            sim_files = max_sim_files
            group.add (max_master)

            masters.remove (max_master)

        if len (group) < group_len:
            continue
        l_group = list (group)
        
        for i in range (len (l_group)):
            sheet[get_column_letter (i+1) + str (row)] = l_group[i]

        sheet[get_column_letter (group_len + 1) + str (row)] = str (sim_files)
        sheet[get_column_letter (group_len + 2) + str (row)] = str (group_weight (group, db))
        
        row += 1









            
